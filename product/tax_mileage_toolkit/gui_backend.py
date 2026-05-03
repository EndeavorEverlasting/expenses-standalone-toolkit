import csv
import json
import re
import threading
from datetime import datetime
from pathlib import Path
from typing import Any

import uvicorn
from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from .audit import run_audit
from .reconcile import _backup_workbook, run_reconcile, run_suggest_clusters
from .reporting import render_html_reports


class RunRequest(BaseModel):
    workbook_path: str
    engage_deferred: bool = False
    write_suggestions: bool = False


class PromoteRequest(BaseModel):
    workbook_path: str
    run_id: str
    row_indices: list[int] = Field(default_factory=list)
    dry_run: bool = True


def create_app(workspace: Path) -> FastAPI:
    app = FastAPI(title="Mileage GUI Workbench")
    gui_dir = Path(__file__).with_name("gui")
    runs_root = workspace / "scripts" / "runs"
    runs_root.mkdir(parents=True, exist_ok=True)
    _lock = threading.Lock()
    _run_state: dict[str, Any] = {"active": False, "last_run_id": ""}
    _run_logs: dict[str, list[str]] = {}
    _run_done: dict[str, bool] = {}
    _run_error: dict[str, str | None] = {}

    app.mount("/static", StaticFiles(directory=gui_dir), name="static")

    def _safe_path(raw: str) -> Path:
        p = Path(raw).expanduser().resolve()
        if not p.is_relative_to(workspace):
            raise HTTPException(status_code=400, detail="Path outside workspace is not allowed.")
        return p

    def _get_run_dir(run_id: str) -> Path:
        if not re.fullmatch(r"[A-Za-z0-9_-]+", run_id or ""):
            raise HTTPException(status_code=400, detail="Invalid run id.")
        run_dir = (runs_root / run_id).resolve()
        if not run_dir.is_relative_to(runs_root):
            raise HTTPException(status_code=400, detail="Invalid run id.")
        if not run_dir.exists():
            raise HTTPException(status_code=404, detail="Run not found.")
        return run_dir

    def _list_runs() -> list[dict[str, Any]]:
        if not runs_root.exists():
            return []
        rows = []
        for p in sorted([d for d in runs_root.iterdir() if d.is_dir()], reverse=True):
            rows.append(
                {
                    "id": p.name,
                    "path": str(p),
                    "has_audit": (p / "audit_report.json").exists(),
                    "has_suggestions": (p / "cluster_suggestion_report.csv").exists(),
                    "has_matches": (p / "cluster_match_report.csv").exists(),
                    "has_overlaps": (p / "cluster_overlap_report.csv").exists(),
                    "has_index_html": (p / "index.html").exists(),
                    "has_index_suite_html": (p / "index_suite.html").exists(),
                }
            )
        return rows

    def _read_csv(path: Path) -> list[dict[str, str]]:
        if not path.exists():
            return []
        with path.open("r", encoding="utf-8", newline="") as f:
            return list(csv.DictReader(f))

    def _legacy_actionable_feedback(audit: dict[str, Any]) -> list[dict[str, Any]]:
        defaults = {
            "clusters_missing_final_site_decision": {
                "sheet": "Coord Reconcile Console",
                "columns": ["M (final_site_decision)"],
                "row_start": 6,
                "practical_action": "Fill final_site_decision for unresolved clusters.",
                "alignment_goal": "Finalize cluster decisions before promotion.",
            },
            "known_sites_not_user_confirmed": {
                "sheet": "Known Site Rollup",
                "columns": ["M (user_confirmed)"],
                "row_start": 6,
                "practical_action": "Confirm each known site and set user_confirmed.",
                "alignment_goal": "Keep matching restricted to confirmed sites.",
            },
            "distance_matrix_active_rows_missing_standard_miles": {
                "sheet": "Sites & Distance Matrix",
                "columns": ["D (active_pair)", "K (standard_miles)"],
                "row_start": 6,
                "practical_action": "Enter standard_miles for active route pairs.",
                "alignment_goal": "Ensure consistent baseline mileage values.",
            },
            "site_day_rows_missing_final_site_decision": {
                "sheet": "Site-Day Draft",
                "columns": ["S (final_site_decision)"],
                "row_start": 5,
                "practical_action": "Set final_site_decision for unresolved day rows.",
                "alignment_goal": "Align day rows with finalized sites.",
            },
            "weekend_site_day_rows_pending_review": {
                "sheet": "Site-Day Draft",
                "columns": ["P (day_type)", "T (weekend_review_status)"],
                "row_start": 5,
                "practical_action": "Review weekend rows and update review status.",
                "alignment_goal": "Separate deductible vs non-deductible weekend travel.",
            },
            "drafted_legs_missing_classification": {
                "sheet": "Mileage Detail - Drafted",
                "columns": ["B (leg_identifier)", "O (classification)"],
                "row_start": 5,
                "practical_action": "Classify each drafted leg.",
                "alignment_goal": "Make totals and downstream filtering accurate.",
            },
            "toll_candidate_legs_missing_toll_decision": {
                "sheet": "Mileage Detail - Drafted",
                "columns": ["V (toll_candidate)", "W (toll_decision)"],
                "row_start": 5,
                "practical_action": "Record toll decisions for toll candidate legs.",
                "alignment_goal": "Include only valid toll charges in totals.",
            },
        }
        feedback = []
        for key, template in defaults.items():
            count = audit.get(key, 0)
            if not isinstance(count, int):
                continue
            feedback.append(
                {
                    "metric_key": key,
                    "count": count,
                    "workbook_location": {
                        "sheet": template["sheet"],
                        "columns": template["columns"],
                        "row_start": template["row_start"],
                        "row_end": None,
                    },
                    "practical_action": template["practical_action"],
                    "alignment_goal": template["alignment_goal"],
                    "sample_rows": [],
                }
            )
        return feedback

    def _append_log(run_id: str, message: str) -> None:
        ts = datetime.now().strftime("%H:%M:%S")
        _run_logs[run_id].append(f"[{ts}] {message}")

    def _execute_run(run_id: str, run_dir: Path, workbook_path: Path, payload: RunRequest) -> None:
        try:
            _append_log(run_id, f"Run {run_id} started.")
            _append_log(run_id, "Step 1/4 — Auditing workbook…")
            audit = run_audit(workbook_path, run_dir)
            _append_log(run_id, "Audit complete.")

            _append_log(run_id, "Step 2/4 — Reconciling cluster distances…")
            reconcile = run_reconcile(workbook_path, run_dir)
            _append_log(run_id, "Reconcile complete.")

            _append_log(run_id, "Step 3/4 — Suggesting cluster matches…")
            suggest = run_suggest_clusters(
                workbook_path,
                run_dir,
                dry_run=not payload.write_suggestions,
                engage_deferred=payload.engage_deferred,
            )
            suggested_count = suggest.get("suggested", 0) if isinstance(suggest, dict) else 0
            _append_log(run_id, f"Suggestions complete — {suggested_count} suggestion(s) generated.")

            _append_log(run_id, "Step 4/4 — Rendering HTML reports…")
            html = render_html_reports(run_dir)
            _append_log(run_id, "HTML reports ready.")

            result = {
                "run_id": run_id,
                "run_dir": str(run_dir),
                "audit": audit,
                "reconcile": reconcile,
                "suggest": suggest,
                "html": html,
            }
            (run_dir / "run_result.json").write_text(json.dumps(result, default=str), encoding="utf-8")
            _run_state["last_run_id"] = run_id
            _append_log(run_id, "Run complete.")
            _run_error[run_id] = None
        except Exception as exc:
            _append_log(run_id, f"Error: {exc}")
            _run_error[run_id] = str(exc)
        finally:
            _run_done[run_id] = True
            _run_state["active"] = False
            _prune_run_logs(keep=20)

    def _prune_run_logs(keep: int = 20) -> None:
        keys = list(_run_logs.keys())
        for old_key in keys[:-keep]:
            _run_logs.pop(old_key, None)
            _run_done.pop(old_key, None)
            _run_error.pop(old_key, None)

    @app.get("/api/browse")
    def api_browse() -> dict[str, Any]:
        import subprocess
        import sys

        _dialog_script = (
            "import sys, tkinter as tk; from tkinter import filedialog; "
            "initialdir = sys.argv[1] if len(sys.argv) > 1 else '.'; "
            "root = tk.Tk(); root.withdraw(); root.wm_attributes('-topmost', True); "
            "path = filedialog.askopenfilename("
            "title='Select Workbook', "
            "initialdir=initialdir, "
            "filetypes=[('Excel files', '*.xlsx *.xls')]"
            "); root.destroy(); print(path or '', end='')"
        )
        try:
            result = subprocess.run(
                [sys.executable, "-c", _dialog_script, str(workspace)],
                capture_output=True,
                text=True,
                timeout=120,
            )
            if result.returncode != 0:
                err = result.stderr.strip() or "unknown error"
                raise HTTPException(status_code=500, detail=f"File dialog failed: {err}")
            selected = result.stdout.strip()
            if selected and not selected.lower().endswith((".xlsx", ".xls")):
                raise HTTPException(status_code=400, detail="Selected file must be an .xlsx or .xls workbook.")
            return {"path": selected}
        except subprocess.TimeoutExpired:
            raise HTTPException(status_code=408, detail="File dialog timed out.")
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"File dialog failed: {exc}")

    @app.get("/")
    def root() -> FileResponse:
        return FileResponse(gui_dir / "index.html")

    @app.get("/api/runs")
    def api_runs() -> dict[str, Any]:
        return {"runs": _list_runs(), "state": _run_state}

    @app.get("/api/runs/{run_id}/summary")
    def api_summary(run_id: str) -> dict[str, Any]:
        run_dir = _get_run_dir(run_id)

        audit = {}
        audit_path = run_dir / "audit_report.json"
        if audit_path.exists():
            audit = json.loads(audit_path.read_text(encoding="utf-8"))
        actionable_feedback = audit.get("actionable_feedback")
        if not isinstance(actionable_feedback, list):
            actionable_feedback = _legacy_actionable_feedback(audit)

        suggestions = _read_csv(run_dir / "cluster_suggestion_report.csv")
        matches = _read_csv(run_dir / "cluster_match_report.csv")
        overlaps = _read_csv(run_dir / "cluster_overlap_report.csv")
        suggested = sum(1 for r in suggestions if r.get("status") == "suggested")
        deferred = sum(1 for r in suggestions if r.get("status") == "deferred")
        skipped = sum(1 for r in suggestions if r.get("status") == "skipped")
        return {
            "run_id": run_id,
            "audit": audit,
            "actionable_feedback": actionable_feedback,
            "suggestions": {"total": len(suggestions), "suggested": suggested, "deferred": deferred, "skipped": skipped},
            "matches_rows": len(matches),
            "overlaps_rows": len(overlaps),
        }

    @app.get("/api/runs/{run_id}/log")
    def api_run_log(run_id: str, since: int = 0) -> dict[str, Any]:
        if not re.fullmatch(r"[A-Za-z0-9_-]+", run_id or ""):
            raise HTTPException(status_code=400, detail="Invalid run id.")
        logs = _run_logs.get(run_id)
        if logs is None:
            raise HTTPException(status_code=404, detail="No log found for this run id.")
        since = max(0, since)
        return {
            "lines": logs[since:],
            "total": len(logs),
            "done": _run_done.get(run_id, True),
            "error": _run_error.get(run_id),
        }

    @app.get("/api/runs/{run_id}/cluster-stats")
    def api_cluster_stats(run_id: str) -> dict[str, Any]:
        run_dir = _get_run_dir(run_id)
        matches = _read_csv(run_dir / "cluster_match_report.csv")
        overlaps = _read_csv(run_dir / "cluster_overlap_report.csv")

        total_clusters = len(matches)
        total_distinct_days = 0
        resolved_count = 0
        unresolved_count = 0
        for row in matches:
            try:
                total_distinct_days += int(row.get("distinct_days") or 0)
            except (ValueError, TypeError):
                pass
            status = (row.get("review_status") or "").strip().lower()
            if status and status not in ("unresolved", "pending", ""):
                resolved_count += 1
            else:
                unresolved_count += 1

        grade_counts: dict[str, int] = {"Strong": 0, "Near": 0, "unmatched": 0}
        for row in matches:
            grade = (row.get("auto_match_grade") or "").strip()
            if grade == "Strong":
                grade_counts["Strong"] += 1
            elif grade == "Near":
                grade_counts["Near"] += 1
            else:
                grade_counts["unmatched"] += 1

        return {
            "run_id": run_id,
            "total_clusters": total_clusters,
            "total_distinct_days": total_distinct_days,
            "resolved_count": resolved_count,
            "unresolved_count": unresolved_count,
            "overlap_pairs": len(overlaps),
            "grade_counts": grade_counts,
        }

    @app.get("/api/runs/{run_id}/table/{name}")
    def api_table(run_id: str, name: str, q: str = "", limit: int = 500) -> dict[str, Any]:
        run_dir = _get_run_dir(run_id)
        mapping = {
            "suggestions": run_dir / "cluster_suggestion_report.csv",
            "matches": run_dir / "cluster_match_report.csv",
            "overlaps": run_dir / "cluster_overlap_report.csv",
        }
        if name not in mapping:
            raise HTTPException(status_code=400, detail="Unknown table name.")
        rows = _read_csv(mapping[name])
        if q:
            ql = q.lower()
            rows = [r for r in rows if ql in json.dumps(r).lower()]
        rows = rows[: max(1, min(limit, 5000))]
        return {"rows": rows, "total": len(rows)}

    @app.post("/api/run")
    def api_run(payload: RunRequest) -> dict[str, Any]:
        workbook_path = _safe_path(payload.workbook_path)
        if not workbook_path.exists():
            raise HTTPException(status_code=404, detail="Workbook path not found.")

        with _lock:
            if _run_state["active"]:
                raise HTTPException(status_code=409, detail="Another run is currently active.")
            _run_state["active"] = True
            run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
            run_dir = runs_root / run_id
            run_dir.mkdir(parents=True, exist_ok=True)
            _run_logs[run_id] = []
            _run_done[run_id] = False
            _run_error[run_id] = None

        t = threading.Thread(
            target=_execute_run,
            args=(run_id, run_dir, workbook_path, payload),
            daemon=True,
        )
        t.start()
        return {"run_id": run_id, "status": "running"}

    @app.post("/api/promote")
    def api_promote(payload: PromoteRequest) -> dict[str, Any]:
        workbook_path = _safe_path(payload.workbook_path)
        run_dir = _get_run_dir(payload.run_id)
        if not workbook_path.exists():
            raise HTTPException(status_code=404, detail="Workbook not found.")

        selected_rows = set(payload.row_indices)
        if not selected_rows:
            raise HTTPException(status_code=400, detail="row_indices must include at least one row.")

        suggestions = _read_csv(run_dir / "cluster_suggestion_report.csv")
        suggestion_map = {
            int(r["row_idx"]): r
            for r in suggestions
            if r.get("status") == "suggested" and r.get("nearest_site")
        }
        apply_rows = [idx for idx in selected_rows if idx in suggestion_map]
        skipped_rows = [idx for idx in selected_rows if idx not in suggestion_map]

        report = {
            "dry_run": payload.dry_run,
            "requested_rows": sorted(selected_rows),
            "eligible_rows": sorted(apply_rows),
            "skipped_rows": sorted(skipped_rows),
            "applied_count": 0,
            "backup_path": "",
        }
        if not payload.dry_run and apply_rows:
            from openpyxl import load_workbook

            backup_path = _backup_workbook(workbook_path, run_dir / "promotion_backups")
            wb = load_workbook(workbook_path)
            ws = wb["Maps 2025 - Clusters"]
            for row_idx in apply_rows:
                ws.cell(row_idx, 25).value = suggestion_map[row_idx]["nearest_site"]
                report["applied_count"] += 1
            wb.save(workbook_path)
            report["backup_path"] = str(backup_path)

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = run_dir / f"promotion_report_{stamp}.json"
        out.write_text(json.dumps(report, indent=2), encoding="utf-8")
        report["promotion_report"] = str(out)
        return report

    return app


def serve_gui(host: str = "0.0.0.0", port: int = 5000, workspace: Path | None = None) -> None:
    root = workspace.expanduser().resolve() if workspace else Path(__file__).resolve().parents[2]
    app = create_app(root)
    uvicorn.run(app, host=host, port=port)
