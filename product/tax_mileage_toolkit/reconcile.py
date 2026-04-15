import csv
import math
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

EARTH_RADIUS_MILES = 3958.7613


def haversine_miles(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * EARTH_RADIUS_MILES * math.asin(math.sqrt(a))


def _load_known_sites(ws: Any) -> list[dict[str, Any]]:
    rows = []
    for r in range(11, ws.max_row + 1):
        name = ws.cell(r, 2).value
        lat = ws.cell(r, 6).value
        lng = ws.cell(r, 7).value
        if name and lat not in (None, "") and lng not in (None, ""):
            rows.append(
                {
                    "site_id": ws.cell(r, 1).value,
                    "site_name": name,
                    "address": ws.cell(r, 5).value,
                    "lat": float(lat),
                    "lng": float(lng),
                    "toll_region": ws.cell(r, 15).value,
                }
            )
    return rows


def _load_clusters(ws: Any) -> list[dict[str, Any]]:
    rows = []
    for r in range(6, ws.max_row + 1):
        cid = ws.cell(r, 1).value
        lat = ws.cell(r, 11).value
        lng = ws.cell(r, 12).value
        if cid and lat not in (None, "") and lng not in (None, ""):
            rows.append(
                {
                    "cluster_id": cid,
                    "review_status": ws.cell(r, 2).value,
                    "user_site_label": ws.cell(r, 3).value,
                    "candidate_work_site": ws.cell(r, 4).value,
                    "distinct_days": ws.cell(r, 7).value,
                    "first_seen": ws.cell(r, 8).value,
                    "last_seen": ws.cell(r, 9).value,
                    "months_seen": ws.cell(r, 10).value,
                    "lat": float(lat),
                    "lng": float(lng),
                    "total_hours": ws.cell(r, 14).value,
                }
            )
    return rows


def _load_clusters_with_rows(ws: Any) -> list[dict[str, Any]]:
    rows = []
    for r in range(6, ws.max_row + 1):
        cid = ws.cell(r, 1).value
        lat = ws.cell(r, 11).value
        lng = ws.cell(r, 12).value
        if cid and lat not in (None, "") and lng not in (None, ""):
            rows.append(
                {
                    "row_idx": r,
                    "cluster_id": cid,
                    "review_status": ws.cell(r, 2).value,
                    "user_site_label": ws.cell(r, 3).value,
                    "candidate_work_site": ws.cell(r, 4).value,
                    "final_site_decision": ws.cell(r, 25).value,
                    "lat": float(lat),
                    "lng": float(lng),
                }
            )
    return rows


def _get_thresholds(ws: Any) -> tuple[float, float, float]:
    strong = float(ws["B5"].value or 0.5)
    near = float(ws["B6"].value or 1.5)
    overlap = float(ws["B7"].value or 0.5)
    return strong, near, overlap


def _export_cluster_matches(
    out_dir: Path, clusters: list[dict[str, Any]], known_sites: list[dict[str, Any]], strong: float, near: float
) -> list[dict[str, Any]]:
    fieldnames = [
        "cluster_id",
        "review_status",
        "user_site_label",
        "candidate_work_site",
        "distinct_days",
        "first_seen",
        "last_seen",
        "lat",
        "lng",
        "nearest_site",
        "distance_to_site_mi",
        "auto_match_grade",
        "nearby_sites_le_1_5mi",
        "toll_region",
    ]
    rows = []
    if not clusters:
        out_path = out_dir / "cluster_match_report.csv"
        with out_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
        return rows

    for cluster in clusters:
        distances = []
        for site in known_sites:
            miles = haversine_miles(cluster["lat"], cluster["lng"], site["lat"], site["lng"])
            distances.append((miles, site))
        distances.sort(key=lambda x: x[0])
        if distances:
            best_miles, best_site = distances[0]
            nearby = [f'{site["site_name"]} ({miles:.2f} mi)' for miles, site in distances if miles <= near]
            grade = "Strong" if best_miles <= strong else ("Near" if best_miles <= near else "")
            nearest_site = best_site["site_name"]
            toll_region = best_site.get("toll_region") or ""
            distance_to_site_mi: float | None = round(best_miles, 3)
            nearby_sites_le_1_5mi = "; ".join(nearby)
        else:
            nearest_site = ""
            distance_to_site_mi = None
            grade = ""
            nearby_sites_le_1_5mi = ""
            toll_region = ""
        rows.append(
            {
                "cluster_id": cluster["cluster_id"],
                "review_status": cluster["review_status"],
                "user_site_label": cluster["user_site_label"],
                "candidate_work_site": cluster["candidate_work_site"],
                "distinct_days": cluster["distinct_days"],
                "first_seen": cluster["first_seen"],
                "last_seen": cluster["last_seen"],
                "lat": cluster["lat"],
                "lng": cluster["lng"],
                "nearest_site": nearest_site,
                "distance_to_site_mi": distance_to_site_mi,
                "auto_match_grade": grade,
                "nearby_sites_le_1_5mi": nearby_sites_le_1_5mi,
                "toll_region": toll_region,
            }
        )
    out_path = out_dir / "cluster_match_report.csv"
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return rows


def _export_cluster_overlaps(out_dir: Path, clusters: list[dict[str, Any]], overlap: float) -> list[dict[str, Any]]:
    rows = []
    for i, c1 in enumerate(clusters):
        for c2 in clusters[i + 1 :]:
            miles = haversine_miles(c1["lat"], c1["lng"], c2["lat"], c2["lng"])
            if miles <= overlap:
                rows.append(
                    {
                        "cluster_id_1": c1["cluster_id"],
                        "cluster_id_2": c2["cluster_id"],
                        "distance_mi": round(miles, 3),
                        "days_1": c1["distinct_days"],
                        "days_2": c2["distinct_days"],
                        "user_label_1": c1["user_site_label"] or "",
                        "user_label_2": c2["user_site_label"] or "",
                    }
                )
    out_path = out_dir / "cluster_overlap_report.csv"
    with out_path.open("w", newline="", encoding="utf-8") as f:
        if rows:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
        else:
            f.write("cluster_id_1,cluster_id_2,distance_mi,days_1,days_2,user_label_1,user_label_2\n")
    return rows


def _export_known_site_rollup(
    out_dir: Path, known_sites: list[dict[str, Any]], cluster_rows: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    fieldnames = [
        "site_id",
        "site_name",
        "address",
        "lat",
        "lng",
        "strong_cluster_count",
        "near_cluster_count",
        "toll_region",
    ]
    counts = {site["site_name"]: {"strong": 0, "near": 0} for site in known_sites}
    for row in cluster_rows:
        site = row["nearest_site"]
        if not site:
            continue
        if row["auto_match_grade"] == "Strong":
            counts[site]["strong"] += 1
        elif row["auto_match_grade"] == "Near":
            counts[site]["near"] += 1
    rows = []
    for site in known_sites:
        rows.append(
            {
                "site_id": site["site_id"],
                "site_name": site["site_name"],
                "address": site["address"],
                "lat": site["lat"],
                "lng": site["lng"],
                "strong_cluster_count": counts[site["site_name"]]["strong"],
                "near_cluster_count": counts[site["site_name"]]["near"],
                "toll_region": site["toll_region"] or "",
            }
        )
    out_path = out_dir / "known_site_rollup_report.csv"
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        if rows:
            writer.writerows(rows)
    return rows


def run_reconcile(workbook_path: Path, output_dir: Path | None = None) -> dict[str, int]:
    workbook_path = workbook_path.expanduser().resolve()
    out_dir = output_dir.expanduser().resolve() if output_dir else workbook_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(workbook_path, data_only=True)
    strong, near, overlap = _get_thresholds(wb["Known Site Registry"])
    known_sites = _load_known_sites(wb["Known Site Registry"])
    clusters = _load_clusters(wb["Maps 2025 - Clusters"])

    cluster_rows = _export_cluster_matches(out_dir, clusters, known_sites, strong, near)
    overlap_rows = _export_cluster_overlaps(out_dir, clusters, overlap)
    site_rows = _export_known_site_rollup(out_dir, known_sites, cluster_rows)

    return {
        "cluster_match_report_rows": len(cluster_rows),
        "cluster_overlap_report_rows": len(overlap_rows),
        "known_site_rollup_report_rows": len(site_rows),
    }


def _helper_columns(ws: Any) -> dict[str, int]:
    header_row = 5
    wanted = [
        "suggested_site",
        "suggestion_confidence",
        "suggestion_distance_mi",
        "suggestion_reason",
        "suggestion_run_timestamp",
    ]
    existing: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        value = ws.cell(header_row, c).value
        if value:
            existing[str(value).strip()] = c

    col = ws.max_column + 1
    for name in wanted:
        if name not in existing:
            ws.cell(header_row, col).value = name
            existing[name] = col
            col += 1
    return {k: existing[k] for k in wanted}


def _backup_workbook(workbook_path: Path, backup_dir: Path) -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{workbook_path.stem}.backup_{stamp}{workbook_path.suffix}"
    shutil.copy2(workbook_path, backup_path)
    return backup_path


def _cluster_suggestion(
    cluster: dict[str, Any],
    known_sites: list[dict[str, Any]],
    strong: float,
    near: float,
    ambiguity_margin: float,
    max_suggestion_distance: float,
) -> dict[str, Any]:
    if cluster["final_site_decision"] not in (None, ""):
        return {"status": "skipped", "reason": "already_finalized"}

    distances = []
    for site in known_sites:
        miles = haversine_miles(cluster["lat"], cluster["lng"], site["lat"], site["lng"])
        distances.append((miles, site))
    if not distances:
        return {"status": "deferred", "reason": "missing_site_data"}

    distances.sort(key=lambda x: x[0])
    best_miles, best_site = distances[0]
    close_competitors = [d for d in distances[1:] if d[0] <= near and abs(d[0] - best_miles) <= ambiguity_margin]

    if best_miles > max_suggestion_distance:
        return {
            "status": "deferred",
            "reason": "outside_suggestion_radius",
            "distance_to_site_mi": round(best_miles, 3),
            "nearest_site": best_site["site_name"],
        }

    if close_competitors:
        return {
            "status": "deferred",
            "reason": "ambiguous",
            "distance_to_site_mi": round(best_miles, 3),
            "nearest_site": best_site["site_name"],
        }

    return {
        "status": "suggested",
        "reason": "strong_unambiguous" if best_miles <= strong else "near_unambiguous",
        "nearest_site": best_site["site_name"],
        "distance_to_site_mi": round(best_miles, 3),
        "confidence": "High" if best_miles <= strong else "Medium",
    }


def run_suggest_clusters(
    workbook_path: Path,
    output_dir: Path | None = None,
    dry_run: bool = True,
    ambiguity_margin: float = 0.1,
    engage_deferred: bool = False,
) -> dict[str, Any]:
    workbook_path = workbook_path.expanduser().resolve()
    out_dir = output_dir.expanduser().resolve() if output_dir else workbook_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    wb_data = load_workbook(workbook_path, data_only=True)
    strong, near, _ = _get_thresholds(wb_data["Known Site Registry"])
    known_sites = _load_known_sites(wb_data["Known Site Registry"])
    clusters = _load_clusters_with_rows(wb_data["Maps 2025 - Clusters"])

    stamp = datetime.now().isoformat(timespec="seconds")
    max_suggestion_distance = near if engage_deferred else strong
    rows: list[dict[str, Any]] = []
    for cluster in clusters:
        result = _cluster_suggestion(
            cluster,
            known_sites,
            strong,
            near,
            ambiguity_margin,
            max_suggestion_distance=max_suggestion_distance,
        )
        rows.append(
            {
                "row_idx": cluster["row_idx"],
                "cluster_id": cluster["cluster_id"],
                "nearest_site": result.get("nearest_site", ""),
                "distance_to_site_mi": result.get("distance_to_site_mi", ""),
                "confidence": result.get("confidence", ""),
                "status": result["status"],
                "reason": result["reason"],
                "run_timestamp": stamp,
            }
        )

    suggested_count = sum(1 for r in rows if r["status"] == "suggested")
    deferred_count = sum(1 for r in rows if r["status"] == "deferred")
    skipped_count = sum(1 for r in rows if r["status"] == "skipped")

    backup_path = None
    updated_rows = 0
    if not dry_run:
        backup_path = _backup_workbook(workbook_path, out_dir / "backups")
        wb_write = load_workbook(workbook_path)
        console_ws = wb_write["Coord Reconcile Console"]
        cols = _helper_columns(console_ws)
        for row in rows:
            if row["status"] != "suggested":
                continue
            r = row["row_idx"]
            console_ws.cell(r, cols["suggested_site"]).value = row["nearest_site"]
            console_ws.cell(r, cols["suggestion_confidence"]).value = row["confidence"]
            console_ws.cell(r, cols["suggestion_distance_mi"]).value = row["distance_to_site_mi"]
            console_ws.cell(r, cols["suggestion_reason"]).value = row["reason"]
            console_ws.cell(r, cols["suggestion_run_timestamp"]).value = row["run_timestamp"]
            updated_rows += 1
        wb_write.save(workbook_path)

    csv_path = out_dir / "cluster_suggestion_report.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "row_idx",
            "cluster_id",
            "nearest_site",
            "distance_to_site_mi",
            "confidence",
            "status",
            "reason",
            "run_timestamp",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        if rows:
            writer.writerows(rows)

    summary = {
        "dry_run": dry_run,
        "backup_path": str(backup_path) if backup_path else "",
        "report_csv": str(csv_path),
        "total_clusters_evaluated": len(rows),
        "helper_rows_suggested": suggested_count,
        "rows_deferred": deferred_count,
        "rows_skipped_finalized": skipped_count,
        "helper_rows_written": updated_rows,
        "engage_deferred": engage_deferred,
    }
    return summary
