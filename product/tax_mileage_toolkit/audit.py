import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _feedback_item(
    metric_key: str,
    count: int,
    sheet: str,
    columns: list[str],
    row_start: int,
    row_end: int,
    practical_action: str,
    alignment_goal: str,
    sample_rows: list[int],
) -> dict[str, Any]:
    return {
        "metric_key": metric_key,
        "count": count,
        "workbook_location": {
            "sheet": sheet,
            "columns": columns,
            "row_start": row_start,
            "row_end": row_end,
        },
        "practical_action": practical_action,
        "alignment_goal": alignment_goal,
        "sample_rows": sample_rows[:5],
    }


def run_audit(workbook_path: Path, output_dir: Path | None = None) -> dict[str, Any]:
    workbook_path = workbook_path.expanduser().resolve()
    out_dir = output_dir.expanduser().resolve() if output_dir else workbook_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(workbook_path, data_only=True)

    coord = wb["Coord Reconcile Console"]
    site_roll = wb["Known Site Rollup"]
    matrix = wb["Sites & Distance Matrix"]
    site_day = wb["Site-Day Draft"]
    drafted = wb["Mileage Detail - Drafted"]

    cluster_rows_missing_final = [r for r in range(6, coord.max_row + 1) if coord.cell(r, 13).value in (None, "")]
    clusters_missing_final = len(cluster_rows_missing_final)
    site_rows_unconfirmed = [r for r in range(6, site_roll.max_row + 1) if site_roll.cell(r, 13).value in (None, "")]
    sites_unconfirmed = len(site_rows_unconfirmed)

    matrix_missing_miles = 0
    matrix_rows_missing_miles: list[int] = []
    for r in range(6, matrix.max_row + 1):
        if matrix.cell(r, 4).value == "Yes" and matrix.cell(r, 10).value not in (None, "") and matrix.cell(
            r, 11
        ).value in (None, ""):
            matrix_missing_miles += 1
            matrix_rows_missing_miles.append(r)

    site_day_rows_missing_final = [r for r in range(5, site_day.max_row + 1) if site_day.cell(r, 19).value in (None, "")]
    site_day_missing_final = len(site_day_rows_missing_final)

    weekend_pending = 0
    weekend_pending_rows: list[int] = []
    for r in range(5, site_day.max_row + 1):
        if site_day.cell(r, 16).value == "Weekend" and site_day.cell(r, 20).value in (None, ""):
            weekend_pending += 1
            weekend_pending_rows.append(r)

    drafted_missing_classification = 0
    drafted_toll_missing = 0
    drafted_rows_missing_classification: list[int] = []
    drafted_rows_toll_missing: list[int] = []
    for r in range(5, drafted.max_row + 1):
        if drafted.cell(r, 2).value not in (None, "") and drafted.cell(r, 15).value in (None, ""):
            drafted_missing_classification += 1
            drafted_rows_missing_classification.append(r)
        if drafted.cell(r, 22).value == "Yes" and drafted.cell(r, 23).value in (None, ""):
            drafted_toll_missing += 1
            drafted_rows_toll_missing.append(r)

    actionable_feedback = [
        _feedback_item(
            metric_key="clusters_missing_final_site_decision",
            count=clusters_missing_final,
            sheet="Coord Reconcile Console",
            columns=["M (final_site_decision)"],
            row_start=6,
            row_end=coord.max_row,
            practical_action="Open these rows, confirm the correct work site for each unresolved cluster, and fill final_site_decision.",
            alignment_goal="Cluster-level decisions must be finalized so downstream site-day and mileage rows align to one approved site.",
            sample_rows=cluster_rows_missing_final,
        ),
        _feedback_item(
            metric_key="known_sites_not_user_confirmed",
            count=sites_unconfirmed,
            sheet="Known Site Rollup",
            columns=["M (user_confirmed)"],
            row_start=6,
            row_end=site_roll.max_row,
            practical_action="Review site rollup rows, verify each known site is valid for the period, then set user_confirmed.",
            alignment_goal="Only confirmed sites should drive site matching and distance lookups.",
            sample_rows=site_rows_unconfirmed,
        ),
        _feedback_item(
            metric_key="distance_matrix_active_rows_missing_standard_miles",
            count=matrix_missing_miles,
            sheet="Sites & Distance Matrix",
            columns=["D (active_pair)", "J (route_reference)", "K (standard_miles)"],
            row_start=6,
            row_end=matrix.max_row,
            practical_action="For active site pairs that already have a route reference, enter missing standard_miles so commute math is complete.",
            alignment_goal="Active route pairs need baseline miles so workbook calculations line up consistently across days.",
            sample_rows=matrix_rows_missing_miles,
        ),
        _feedback_item(
            metric_key="site_day_rows_missing_final_site_decision",
            count=site_day_missing_final,
            sheet="Site-Day Draft",
            columns=["S (final_site_decision)"],
            row_start=5,
            row_end=site_day.max_row,
            practical_action="Locate each listed day row and set final_site_decision to the site actually worked that day.",
            alignment_goal="Every site-day row needs a final site so mileage detail rows inherit the right destination.",
            sample_rows=site_day_rows_missing_final,
        ),
        _feedback_item(
            metric_key="weekend_site_day_rows_pending_review",
            count=weekend_pending,
            sheet="Site-Day Draft",
            columns=["P (day_type)", "T (weekend_review_status)"],
            row_start=5,
            row_end=site_day.max_row,
            practical_action="For weekend rows, review supporting notes/calendar evidence and mark weekend_review_status.",
            alignment_goal="Weekend entries must be explicitly reviewed to avoid misclassifying personal vs deductible travel.",
            sample_rows=weekend_pending_rows,
        ),
        _feedback_item(
            metric_key="drafted_legs_missing_classification",
            count=drafted_missing_classification,
            sheet="Mileage Detail - Drafted",
            columns=["B (leg_identifier)", "O (classification)"],
            row_start=5,
            row_end=drafted.max_row,
            practical_action="Classify each drafted leg (for example business, commute, personal) so totals and filters report correctly.",
            alignment_goal="All drafted legs need a classification to produce accurate deduction totals.",
            sample_rows=drafted_rows_missing_classification,
        ),
        _feedback_item(
            metric_key="toll_candidate_legs_missing_toll_decision",
            count=drafted_toll_missing,
            sheet="Mileage Detail - Drafted",
            columns=["V (toll_candidate)", "W (toll_decision)"],
            row_start=5,
            row_end=drafted.max_row,
            practical_action="Review each toll candidate leg and decide whether a toll is claimable, then fill toll_decision.",
            alignment_goal="Toll candidate rows need explicit decisions so expense totals include only valid toll charges.",
            sample_rows=drafted_rows_toll_missing,
        ),
    ]

    report = {
        "clusters_missing_final_site_decision": clusters_missing_final,
        "known_sites_not_user_confirmed": sites_unconfirmed,
        "distance_matrix_active_rows_missing_standard_miles": matrix_missing_miles,
        "site_day_rows_missing_final_site_decision": site_day_missing_final,
        "weekend_site_day_rows_pending_review": weekend_pending,
        "drafted_legs_missing_classification": drafted_missing_classification,
        "toll_candidate_legs_missing_toll_decision": drafted_toll_missing,
        "actionable_feedback": actionable_feedback,
    }

    out_path = out_dir / "audit_report.json"
    out_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    return report
